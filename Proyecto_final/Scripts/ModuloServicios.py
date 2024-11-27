import openpyxl
from loguru import logger
import general_functions as gf

class Servicios:
    def __init__(self, config, menu=None, __PBT=None, archivo_excel="Insumos/servicios.xlsx"):
        """Constructor de clase

        Args:
            __config (dict): Diccionario de configuración del proyecto
            __PBT (class) : Clase que contiene métodos para la manipulación de la fuente de información de Servicios
            archivo_excel (str): Ruta del archivo de Excel donde se guardan los servicios.
        """
        self.__config = config
        self.__menu_servicios = menu
        self.__PBT = __PBT
        self.__archivo_excel = archivo_excel
        self.__servicios_adicionales = []
        self.__costo_total_servicios = []
        self.__verificado = False
        self.__historial_actividades = []

        self.cargar_servicios_desde_excel()

    def cargar_servicios_desde_excel(self):
        """Carga los servicios y costos desde el archivo Excel"""
        try:
            wb = openpyxl.load_workbook(self.__archivo_excel)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                servicio, costo = fila
                self.asignar_servicio(servicio, costo)

            logger.info("Servicios cargados desde Excel con éxito.")
        except Exception as e:
            logger.error(f"Error al cargar los servicios desde el archivo Excel: {e}")

    def guardar_servicios_en_excel(self):
        """Guarda los servicios y costos actuales en el archivo Excel"""
        try:
            wb = openpyxl.load_workbook(self.__archivo_excel)
            hoja = wb.active 

            # Limpia la hoja para guardar los nuevos datos
            for fila in hoja.iter_rows(min_row=2, max_col=2):
                for celda in fila:
                    celda.value = None

            for i, servicio in enumerate(self.__servicios_adicionales, start=2):
                hoja.cell(row=i, column=1, value=servicio)  # Servicio
                hoja.cell(row=i, column=2, value=self.__costo_total_servicios[i-2])  # Costo

            wb.save(self.__archivo_excel)
            logger.info("Servicios guardados en el archivo Excel con éxito.")
        except Exception as e:
            logger.error(f"Error al guardar los servicios en el archivo Excel: {e}")

    def mostrar_menu(self):
        eleccion = self.__config["Menu"]["menu_opcion"]["5"]
        gf.mostrar_menu_personalizado(eleccion, self.__menu_servicios)

    def ejecutar_proceso(self):
        opcion_ingresada = input("Ingresa la opción a ejecutar:\n ")
        resultado = self.ejecutar_proceso_servicio(opcion_ingresada)
        return resultado

    def ejecutar_proceso_servicio(self, opcion: str) -> bool:
        """
        Ejecuta la opción seleccionada por el usuario.

        Args:
            opcion (str): Opción seleccionada.

        Returns:
            bool: True si debe continuar, False si debe detenerse.
        """  
        opciones = {
            "1": lambda: print(f"Costo total de los servicios: ${self.calcular_costo_total():.2f}"),
            "2": lambda: self.verificar_servicio() or print("Servicio marcado como verificado."),
            "3": lambda: self.asignar_servicio(input("Nombre del servicio: "), float(input("Costo del servicio: "))),
            "4": lambda: self.eliminar_servicio(input("Nombre del servicio a eliminar: ")) or print("Servicio eliminado correctamente."),
            "5": lambda: print(self.generar_reporte_servicios()),  # Opción para generar reporte
            "0": lambda: False,
        }

        return gf.procesar_opcion(opcion=opcion, opciones=opciones)

    def calcular_costo_total(self):
        """
        Calcula el costo total de los servicios asignados.
        :return: Costo total.
        """
        return sum(self.__costo_total_servicios)

    def verificar_servicio(self):
        """
        Marca los servicios como verificados.
        """
        self.__verificado = True
        logger.info("Servicio verificado correctamente.")

    def asignar_servicio(self, servicio, costo):
        """
        Asigna un nuevo servicio adicional y actualiza el costo total.
        Args:
            servicio (str): Nombre del servicio.
            costo (float): Costo del servicio.
        """
        self.__servicios_adicionales.append(servicio)
        self.__costo_total_servicios.append(costo)
        self.__registrar_actividad(f"Servicio '{servicio}' asignado con costo ${costo:.2f}")
        logger.info(f"Servicio '{servicio}' asignado correctamente. Costo actualizado: {self.__costo_total_servicios}.")

    def eliminar_servicio(self, servicio):
        """
        Elimina un servicio adicional y ajusta el costo total.
        Args:
            servicio (str): Nombre del servicio a eliminar.
        """
        if servicio in self.__servicios_adicionales:
            self.__servicios_adicionales.remove(servicio)
            self.__costo_total_servicios.pop(self.__servicios_adicionales.index(servicio))  # Elimina el costo correspondiente
            self.__registrar_actividad(f"Servicio '{servicio}' eliminado.")
            logger.info(f"Servicio '{servicio}' eliminado correctamente.")
        else:
            logger.warning(f"El servicio '{servicio}' no se encuentra en los servicios adicionales.")

    def __registrar_actividad(self, descripcion):
        """
        Registra una actividad realizada en el sistema.
        
        Args:
            descripcion (str): Descripción de la actividad realizada.
        """
        self.__historial_actividades.append(descripcion)

    def generar_reporte_servicios(self):
        """
        Genera un reporte detallado de los servicios realizados.
        
        Returns:
            str: Reporte en formato de texto.
        """
        reporte = "Reporte de Servicios:\n"
        for actividad in self.__historial_actividades:
            reporte += f"- {actividad}\n"
        
        reporte += f"\nServicios actuales asignados:\n"
        for servicio, costo in zip(self.__servicios_adicionales, self.__costo_total_servicios):
            reporte += f"Servicio: {servicio}, Costo: ${costo:.2f}\n"
        
        reporte += f"\nCosto Total de Servicios: ${self.calcular_costo_total():.2f}"
        logger.info("Reporte de servicios generado exitosamente.")
        return reporte
